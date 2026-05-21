"""Tests for the Phase 1C cycle-scoped router (``api/v1/cycles.py``).

We call the route handlers as plain Python functions, passing the
in-memory ``Session`` and a hand-constructed ``AccessContext``. This
skips the full FastAPI stack but exercises:

  * the access pipeline (F2/F3/F5/F6) — via SuperAdmin bypass for the
    happy paths, and a stub permission-cache for the negative ones,
  * the service-layer plumbing (cycle_service + PO helpers),
  * the response-model shape returned by each endpoint.

Mirrors the ``test_cycle_po_helpers`` / ``test_cycle_service`` fixtures
so this file stays self-contained.
"""
from datetime import date, datetime
from types import SimpleNamespace

import pytest
from fastapi import HTTPException

from app.api.v1.cycles import (
    abandon_cycle_endpoint,
    append_purchase_order,
    close_cycle_endpoint,
    cycle_history,
    export_cycle_xlsx,
    get_cycle_bundle,
    inheritance_preview,
    list_cycle_working_sheet,
    list_cycles,
    start_cycle,
)
from app.schemas.quot_order_cycle import (
    AppendPurchaseOrderRequest,
    CycleCloseRequest,
    CycleStartRequest,
)
from app.services.access_service import AccessContext, LocationAccess


pytestmark = pytest.mark.unit


# ----------------------------------------------------------------------
# Fixtures — sqlite in-memory + a seeded Approved quotation + ctx
# ----------------------------------------------------------------------

@pytest.fixture
def in_memory_session():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    from app.core.database import Base
    # Force registration of every model on Base.metadata.
    from app.models import (  # noqa: F401
        company, customer, customer_classification, contact_type,
        cost_template, communication, cost_point, delivery, dia,
        enquiry, financial_year, item, lifecycle_unlock_audit, location,
        menu, ownership_transfer, quot_activity_log, quot_annexure,
        quot_order_cycle, quot_po_working_sheet, quot_purchase_order,
        quot_viability, quotation, quotation_format, raw_material_cost,
        raw_material_cost_log, role_menu_map, terms_condition, user,
        user_location_map, asset,
    )

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


@pytest.fixture
def seeded_quotation_and_customer(in_memory_session):
    """Approved quotation + its customer. Sqlite FKs are off so we skip
    parent rows the cycle path doesn't need."""
    from app.models.customer import CustomerMaster
    from app.models.quotation import QuotSummary

    db = in_memory_session
    customer = CustomerMaster(
        customerId=1, companyId=1, customerName="ACME Construction",
    )
    db.add(customer)
    quot = QuotSummary(
        quotId=1, companyId=1, customerId=1, quotNo="QUOT-1",
        quotDate=date.today(), status="Approved", versionNo=1,
        createdby=1, createdon=datetime.utcnow(),
    )
    db.add(quot)
    db.flush()
    return quot


def _superadmin_ctx(db) -> AccessContext:
    """SuperAdmin bypasses F2/F3/F5/F6 — perfect for happy-path
    coverage where we don't want the access pipeline to interfere."""
    return AccessContext(
        user_id=1,
        company_id=1,
        role_id=1,
        is_super_admin=True,
        is_company_admin=False,
        visible_user_ids=None,
        location=LocationAccess(bypass=True),
        _db=db,
    )


def _restricted_ctx(db, perms: dict) -> AccessContext:
    """Non-SuperAdmin ctx with a pre-populated permission cache so
    ``has_permission`` returns the booleans the caller wants without
    needing to seed RoleMenuMap rows."""
    fake_perm_row = SimpleNamespace(**perms)
    ctx = AccessContext(
        user_id=1,
        company_id=1,
        role_id=99,
        is_super_admin=False,
        is_company_admin=True,  # bypass F5/F6 — we're only exercising F3 here.
        visible_user_ids=None,
        location=LocationAccess(bypass=True),
        _db=db,
    )
    # Pre-stuff the per-request cache so ``has_permission`` short-circuits.
    ctx._perm_cache["Quotations"] = fake_perm_row
    return ctx


# ----------------------------------------------------------------------
# list_cycles
# ----------------------------------------------------------------------

class TestListCycles:
    def test_returns_empty_when_no_cycles_yet(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        ctx = _superadmin_ctx(in_memory_session)
        result = list_cycles(
            quot_id=seeded_quotation_and_customer.quotId,
            include_abandoned=False,
            db=in_memory_session,
            ctx=ctx,
        )
        assert result.cycles == []

    def test_returns_active_and_complete_by_default(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        first = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        first.status = "Complete"
        second = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        result = list_cycles(
            quot_id=seeded_quotation_and_customer.quotId,
            include_abandoned=False,
            db=in_memory_session,
            ctx=ctx,
        )
        cycle_nos = [c.cycleNo for c in result.cycles]
        assert cycle_nos == [first.cycleNo, second.cycleNo]

    def test_filters_abandoned_unless_included(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        c1 = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        c1.status = "Abandoned"
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        default_result = list_cycles(
            quot_id=seeded_quotation_and_customer.quotId,
            include_abandoned=False, db=in_memory_session, ctx=ctx,
        )
        assert default_result.cycles == []

        included = list_cycles(
            quot_id=seeded_quotation_and_customer.quotId,
            include_abandoned=True, db=in_memory_session, ctx=ctx,
        )
        assert [c.status for c in included.cycles] == ["Abandoned"]


# ----------------------------------------------------------------------
# start_cycle
# ----------------------------------------------------------------------

class TestStartCycle:
    def test_first_cycle_returns_cycle_one_active(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        ctx = _superadmin_ctx(in_memory_session)
        body = CycleStartRequest()
        result = start_cycle(
            quot_id=seeded_quotation_and_customer.quotId,
            body=body, db=in_memory_session, ctx=ctx,
        )
        assert result.cycleNo == 1
        assert result.status == "Active"
        assert result.parentCycleId is None
        # Side-effect: quotation auto-Converts on cycle #1.
        assert seeded_quotation_and_customer.status == "Converted"

    def test_rejects_quotation_in_draft(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        seeded_quotation_and_customer.status = "Draft"
        in_memory_session.flush()
        ctx = _superadmin_ctx(in_memory_session)
        with pytest.raises(HTTPException) as exc:
            start_cycle(
                quot_id=seeded_quotation_and_customer.quotId,
                body=CycleStartRequest(),
                db=in_memory_session, ctx=ctx,
            )
        assert exc.value.status_code == 400

    def test_rejects_without_can_start_new_cycle_permission(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        ctx = _restricted_ctx(in_memory_session, perms={
            "CanRead": True, "CanStartNewCycle": False,
        })
        with pytest.raises(HTTPException) as exc:
            start_cycle(
                quot_id=seeded_quotation_and_customer.quotId,
                body=CycleStartRequest(),
                db=in_memory_session, ctx=ctx,
            )
        assert exc.value.status_code == 403

    def test_notes_pass_through(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        ctx = _superadmin_ctx(in_memory_session)
        result = start_cycle(
            quot_id=seeded_quotation_and_customer.quotId,
            body=CycleStartRequest(notes="Phase 1 release"),
            db=in_memory_session, ctx=ctx,
        )
        assert result.notes == "Phase 1 release"


# ----------------------------------------------------------------------
# get_cycle_bundle
# ----------------------------------------------------------------------

class TestGetCycleBundle:
    def test_empty_cycle_bundle_shape(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        bundle = get_cycle_bundle(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        assert bundle.cycle.quotOrderCycleId == cycle.quotOrderCycleId
        assert bundle.purchaseOrders == []
        assert bundle.workingSheetLineCount == 0
        assert bundle.viabilityId is None
        assert bundle.annexureId is None

    def test_bundle_includes_appended_pos(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        append_purchase_order_to_cycle(
            in_memory_session, cycle,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=date.today(), customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        bundle = get_cycle_bundle(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        assert len(bundle.purchaseOrders) == 1
        assert bundle.purchaseOrders[0].poNo == "PO-1"


# ----------------------------------------------------------------------
# close_cycle / abandon_cycle endpoints
# ----------------------------------------------------------------------

class TestCloseCycleEndpoint:
    def test_close_fails_without_preconditions(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()
        ctx = _superadmin_ctx(in_memory_session)
        with pytest.raises(HTTPException) as exc:
            close_cycle_endpoint(
                quot_id=seeded_quotation_and_customer.quotId,
                cycle_id=cycle.quotOrderCycleId,
                body=CycleCloseRequest(),
                db=in_memory_session, ctx=ctx,
            )
        assert exc.value.status_code == 400
        # Both blockers should be in the message.
        assert "annexure" in exc.value.detail.lower()
        assert "PO" in exc.value.detail

    def test_close_succeeds_with_preconditions(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.models.quot_annexure import QuotAnnexure
        from app.models.quot_purchase_order import QuotPurchaseOrder
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.add(QuotPurchaseOrder(
            companyId=1, quotId=seeded_quotation_and_customer.quotId,
            quotOrderCycleId=cycle.quotOrderCycleId,
            isLOI=False, status="Submitted",
            poNo="PO-1", poDate=date.today(),
            customerId=1, createdby=1,
        ))
        in_memory_session.add(QuotAnnexure(
            companyId=1, quotId=seeded_quotation_and_customer.quotId,
            quotOrderCycleId=cycle.quotOrderCycleId,
            status="Approved", versionNo=1, createdby=1,
        ))
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        result = close_cycle_endpoint(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            body=CycleCloseRequest(reason="all delivered"),
            db=in_memory_session, ctx=ctx,
        )
        assert result.status == "Complete"
        assert "all delivered" in (result.notes or "")


class TestAbandonCycleEndpoint:
    def test_abandons_active_cycle(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        result = abandon_cycle_endpoint(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            body=CycleCloseRequest(reason="customer cancelled"),
            db=in_memory_session, ctx=ctx,
        )
        assert result.status == "Abandoned"
        assert "customer cancelled" in (result.notes or "")

    def test_cannot_abandon_complete_cycle(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        cycle.status = "Complete"
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        with pytest.raises(HTTPException) as exc:
            abandon_cycle_endpoint(
                quot_id=seeded_quotation_and_customer.quotId,
                cycle_id=cycle.quotOrderCycleId,
                body=CycleCloseRequest(),
                db=in_memory_session, ctx=ctx,
            )
        assert exc.value.status_code == 400


# ----------------------------------------------------------------------
# append_purchase_order — permission picks on isLOI
# ----------------------------------------------------------------------

class TestAppendPurchaseOrderEndpoint:
    def _make_body(self, po_no: str = "PO-1", is_loi: bool = False):
        return AppendPurchaseOrderRequest(
            isLOI=is_loi,
            poNo=po_no, poDate=date.today(), customerId=1,
            billingAddressManual="123 Test Lane",
            consigneeAddressManual="Site 7, Industrial Area",
        )

    def test_append_first_po_returns_sequence_one(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()
        ctx = _superadmin_ctx(in_memory_session)
        result = append_purchase_order(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            body=self._make_body(),
            db=in_memory_session, ctx=ctx,
        )
        # Schema → ORM-validated to the response model.
        assert result.poNo == "PO-1"
        assert result.quotId == seeded_quotation_and_customer.quotId

    def test_append_to_closed_cycle_is_409(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        cycle.status = "Complete"
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        with pytest.raises(HTTPException) as exc:
            append_purchase_order(
                quot_id=seeded_quotation_and_customer.quotId,
                cycle_id=cycle.quotOrderCycleId,
                body=self._make_body(),
                db=in_memory_session, ctx=ctx,
            )
        assert exc.value.status_code == 409

    def test_loi_requires_can_capture_loi_not_submit_po(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        """An LOI append should be gated by ``CanCaptureLOI``; a user
        with only ``CanSubmitPO`` must be rejected when isLOI=True."""
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _restricted_ctx(in_memory_session, perms={
            "CanRead": True, "CanSubmitPO": True, "CanCaptureLOI": False,
        })
        with pytest.raises(HTTPException) as exc:
            append_purchase_order(
                quot_id=seeded_quotation_and_customer.quotId,
                cycle_id=cycle.quotOrderCycleId,
                body=self._make_body("LOI-1", is_loi=True),
                db=in_memory_session, ctx=ctx,
            )
        assert exc.value.status_code == 403

    def test_formal_po_requires_can_submit_po_not_loi(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _restricted_ctx(in_memory_session, perms={
            "CanRead": True, "CanCaptureLOI": True, "CanSubmitPO": False,
        })
        with pytest.raises(HTTPException) as exc:
            append_purchase_order(
                quot_id=seeded_quotation_and_customer.quotId,
                cycle_id=cycle.quotOrderCycleId,
                body=self._make_body("PO-1", is_loi=False),
                db=in_memory_session, ctx=ctx,
            )
        assert exc.value.status_code == 403


# ----------------------------------------------------------------------
# Phase 1E — inheritance preview + cycle-scoped FWS + auto-clone
# ----------------------------------------------------------------------

def _seed_working_sheet_row(
    db, *, cycle, po, item_name="TMT 12mm", tpwgst="50000.00",
):
    """Compact factory for a single FWS row attached to ``cycle`` +
    ``po``. Returns the inserted row."""
    from decimal import Decimal
    from app.models.quot_po_working_sheet import QuotPOWorkingSheet
    row = QuotPOWorkingSheet(
        companyId=cycle.companyId,
        quotPOId=po.quotPOId,
        quotOrderCycleId=cycle.quotOrderCycleId,
        itemName=item_name, itemGradeName="Fe550D",
        itemDia="12mm", itemUnit="MT",
        quantity=Decimal("100.00"),
        TPWGST=Decimal(tpwgst),
        createdby=1,
    )
    db.add(row)
    db.flush()
    return row


class TestInheritancePreview:
    def test_returns_none_when_parent_empty(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()
        ctx = _superadmin_ctx(in_memory_session)
        result = inheritance_preview(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        assert result.sourceType == "none"
        assert result.lineCount == 0
        assert result.parentCycleNo == cycle.cycleNo

    def test_returns_working_sheet_when_no_approved_viability(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        po = append_purchase_order_to_cycle(
            in_memory_session, cycle,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )
        _seed_working_sheet_row(in_memory_session, cycle=cycle, po=po)
        _seed_working_sheet_row(in_memory_session, cycle=cycle, po=po,
                                item_name="TMT 16mm")

        ctx = _superadmin_ctx(in_memory_session)
        result = inheritance_preview(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        assert result.sourceType == "working_sheet"
        assert result.lineCount == 2

    def test_prefers_approved_viability_over_working_sheet(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from decimal import Decimal
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody
        from app.models.quot_viability import (
            QuotViabilityLine, QuotViabilitySheet,
        )

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        po = append_purchase_order_to_cycle(
            in_memory_session, cycle,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )
        _seed_working_sheet_row(in_memory_session, cycle=cycle, po=po)

        viab = QuotViabilitySheet(
            companyId=1, quotId=seeded_quotation_and_customer.quotId,
            quotOrderCycleId=cycle.quotOrderCycleId,
            status="Approved", versionNo=1, createdby=1,
        )
        in_memory_session.add(viab)
        in_memory_session.flush()
        in_memory_session.add(QuotViabilityLine(
            companyId=1, viabilityId=viab.viabilityId,
            itemName="TMT 12mm", quantity=Decimal("100"),
            TPWGST=Decimal("52000"), createdby=1,
        ))
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        result = inheritance_preview(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        assert result.sourceType == "viability"
        assert result.lineCount == 1


class TestAutoCloneOnFirstAppend:
    """First append on a child cycle clones the parent's inheritance
    source into the child's FWS. Subsequent appends share the same
    rows (one WS per cycle, CR decision C2)."""

    def test_first_append_clones_parent_ws(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody

        parent = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        parent_po = append_purchase_order_to_cycle(
            in_memory_session, parent,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )
        _seed_working_sheet_row(in_memory_session, cycle=parent, po=parent_po,
                                item_name="A")
        _seed_working_sheet_row(in_memory_session, cycle=parent, po=parent_po,
                                item_name="B")

        # Child cycle — auto-resolves parent in service.
        child = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        assert child.parentCycleId == parent.quotOrderCycleId

        # First append should pull the WS forward into the child cycle.
        append_purchase_order_to_cycle(
            in_memory_session, child,
            QuotPurchaseOrderBody(
                poNo="PO-2", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )

        from app.services.po_working_sheet_service import (
            list_working_sheet_for_cycle,
        )
        child_ws = list_working_sheet_for_cycle(in_memory_session, child)
        names = sorted(r.itemName for r in child_ws)
        assert names == ["A", "B"]

    def test_subsequent_append_does_not_double_clone(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody

        parent = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        parent_po = append_purchase_order_to_cycle(
            in_memory_session, parent,
            QuotPurchaseOrderBody(
                poNo="PO-P", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )
        _seed_working_sheet_row(in_memory_session, cycle=parent, po=parent_po)

        child = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        for n in ("PO-1", "LOI-2", "PO-3"):
            append_purchase_order_to_cycle(
                in_memory_session, child,
                QuotPurchaseOrderBody(
                    poNo=n, poDate=__import__("datetime").date.today(),
                    customerId=1,
                    billingAddressManual="A", consigneeAddressManual="B",
                ),
                user_id=1, is_loi=(n.startswith("LOI")),
            )

        from app.services.po_working_sheet_service import (
            list_working_sheet_for_cycle,
        )
        child_ws = list_working_sheet_for_cycle(in_memory_session, child)
        # Parent had exactly one row → child should have exactly one,
        # not three (one per append).
        assert len(child_ws) == 1

    def test_first_cycle_does_not_clone(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        """Cycle #1 has no parent — append must not attempt to clone."""
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        assert cycle.parentCycleId is None

        append_purchase_order_to_cycle(
            in_memory_session, cycle,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )

        from app.services.po_working_sheet_service import (
            list_working_sheet_for_cycle,
        )
        rows = list_working_sheet_for_cycle(in_memory_session, cycle)
        assert rows == []


class TestListCycleWorkingSheet:
    def test_endpoint_returns_active_rows_for_cycle(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        po = append_purchase_order_to_cycle(
            in_memory_session, cycle,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )
        _seed_working_sheet_row(in_memory_session, cycle=cycle, po=po,
                                item_name="Live row")
        # Inactive row should not appear.
        inactive = _seed_working_sheet_row(
            in_memory_session, cycle=cycle, po=po, item_name="Soft-deleted",
        )
        inactive.isActive = False
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        rows = list_cycle_working_sheet(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        assert len(rows) == 1
        assert rows[0].itemName == "Live row"


# ----------------------------------------------------------------------
# Phase 1F — Cycle History timeline
# ----------------------------------------------------------------------

class TestCycleHistory:
    def test_returns_empty_when_no_cycles(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        ctx = _superadmin_ctx(in_memory_session)
        result = cycle_history(
            quot_id=seeded_quotation_and_customer.quotId,
            db=in_memory_session, ctx=ctx,
        )
        assert result.bundles == []

    def test_includes_every_cycle_with_bundle_shape(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody

        c1 = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        append_purchase_order_to_cycle(
            in_memory_session, c1,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )
        c1.status = "Complete"
        c2 = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        result = cycle_history(
            quot_id=seeded_quotation_and_customer.quotId,
            db=in_memory_session, ctx=ctx,
        )
        # Two cycles, ordered ascending by cycleNo.
        assert [b.cycle.cycleNo for b in result.bundles] == [c1.cycleNo, c2.cycleNo]
        # Bundle 1 has the PO; bundle 2 is empty.
        assert len(result.bundles[0].purchaseOrders) == 1
        assert result.bundles[0].purchaseOrders[0].poNo == "PO-1"
        assert result.bundles[1].purchaseOrders == []

    def test_includes_abandoned_cycles(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.services.cycle_service import start_new_cycle

        c1 = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        c1.status = "Abandoned"
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        result = cycle_history(
            quot_id=seeded_quotation_and_customer.quotId,
            db=in_memory_session, ctx=ctx,
        )
        assert len(result.bundles) == 1
        assert result.bundles[0].cycle.status == "Abandoned"


# ----------------------------------------------------------------------
# Phase 1G — Excel export + activity log codes
# ----------------------------------------------------------------------

class TestExportCycleXlsx:
    def test_empty_cycle_returns_summary_and_working_sheet_only(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        """Cycle with no POs / WS / viability still emits a workbook —
        Summary tab + an empty Working Sheet — never an error."""
        from openpyxl import load_workbook
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        resp = export_cycle_xlsx(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        # Response object with body bytes — load and verify shape.
        from io import BytesIO
        wb = load_workbook(BytesIO(resp.body))
        assert "Summary" in wb.sheetnames
        assert "Working Sheet" in wb.sheetnames
        # No viability → no Viability tab.
        assert "Viability Sheet" not in wb.sheetnames
        # Content-Disposition picks up the cycle no + quot no.
        cd = resp.headers["Content-Disposition"]
        assert f"Cycle-{cycle.cycleNo}" in cd
        assert "QUOT-1" in cd

    def test_filename_uses_quot_id_when_quot_no_missing(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        """Defensive: legacy quotations without ``quotNo`` still get a
        usable filename rather than a stringly-typed ``None``."""
        from app.services.cycle_service import start_new_cycle

        seeded_quotation_and_customer.quotNo = None
        in_memory_session.flush()
        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        resp = export_cycle_xlsx(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        cd = resp.headers["Content-Disposition"]
        assert f"Q-{seeded_quotation_and_customer.quotId}" in cd

    def test_cycle_with_viability_includes_viability_sheet(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from decimal import Decimal
        from openpyxl import load_workbook
        from io import BytesIO
        from app.services.cycle_service import start_new_cycle
        from app.services.purchase_order_service import (
            append_purchase_order_to_cycle,
        )
        from app.schemas.quot_purchase_order import QuotPurchaseOrderBody
        from app.models.quot_viability import (
            QuotViabilityLine, QuotViabilitySheet,
        )

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        append_purchase_order_to_cycle(
            in_memory_session, cycle,
            QuotPurchaseOrderBody(
                poNo="PO-1", poDate=__import__("datetime").date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            user_id=1,
        )

        viab = QuotViabilitySheet(
            companyId=1, quotId=seeded_quotation_and_customer.quotId,
            quotOrderCycleId=cycle.quotOrderCycleId,
            status="Approved", versionNo=1, createdby=1,
        )
        in_memory_session.add(viab)
        in_memory_session.flush()
        in_memory_session.add(QuotViabilityLine(
            companyId=1, viabilityId=viab.viabilityId,
            itemName="TMT 12mm", quantity=Decimal("100"),
            TPWGST=Decimal("52000"), createdby=1,
        ))
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        resp = export_cycle_xlsx(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        wb = load_workbook(BytesIO(resp.body))
        assert "Viability Sheet" in wb.sheetnames

    def test_export_writes_activity_log_row(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.models.quot_activity_log import QuotActivityLog
        from app.services.cycle_log_events import CYCLE_EXPORTED_XLSX
        from app.services.cycle_service import start_new_cycle

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()

        ctx = _superadmin_ctx(in_memory_session)
        export_cycle_xlsx(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            db=in_memory_session, ctx=ctx,
        )
        row = (
            in_memory_session.query(QuotActivityLog)
            .filter(QuotActivityLog.action == CYCLE_EXPORTED_XLSX)
            .first()
        )
        assert row is not None
        assert f"cycle #{cycle.cycleNo}" in (row.details or "")


class TestCycleLogEvents:
    """Sanity: every cycle endpoint emits the centralised constant
    string from ``cycle_log_events``, not an ad-hoc free-text phrase.
    Catches drift if a future refactor reverts to inline strings."""

    def test_start_writes_cycle_started_constant(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.models.quot_activity_log import QuotActivityLog
        from app.services.cycle_log_events import CYCLE_STARTED

        ctx = _superadmin_ctx(in_memory_session)
        start_cycle(
            quot_id=seeded_quotation_and_customer.quotId,
            body=CycleStartRequest(),
            db=in_memory_session, ctx=ctx,
        )
        row = (
            in_memory_session.query(QuotActivityLog)
            .filter(QuotActivityLog.action == CYCLE_STARTED)
            .first()
        )
        assert row is not None

    def test_append_po_writes_po_appended_constant(
        self, in_memory_session, seeded_quotation_and_customer,
    ):
        from app.models.quot_activity_log import QuotActivityLog
        from app.services.cycle_log_events import (
            LOI_APPENDED_TO_CYCLE, PO_APPENDED_TO_CYCLE,
        )
        from app.services.cycle_service import start_new_cycle
        from datetime import date

        cycle = start_new_cycle(
            in_memory_session, seeded_quotation_and_customer, started_by=1,
        )
        in_memory_session.flush()
        ctx = _superadmin_ctx(in_memory_session)

        append_purchase_order(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            body=AppendPurchaseOrderRequest(
                isLOI=False, poNo="PO-X", poDate=date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            db=in_memory_session, ctx=ctx,
        )
        append_purchase_order(
            quot_id=seeded_quotation_and_customer.quotId,
            cycle_id=cycle.quotOrderCycleId,
            body=AppendPurchaseOrderRequest(
                isLOI=True, poNo="LOI-Y", poDate=date.today(),
                customerId=1,
                billingAddressManual="A", consigneeAddressManual="B",
            ),
            db=in_memory_session, ctx=ctx,
        )

        actions = {
            r.action for r in
            in_memory_session.query(QuotActivityLog).all()
        }
        assert PO_APPENDED_TO_CYCLE in actions
        assert LOI_APPENDED_TO_CYCLE in actions
