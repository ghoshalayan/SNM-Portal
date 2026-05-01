from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime

from sqlalchemy.orm import aliased
from app.core.dependencies import get_db, get_current_user, CurrentUser, require_permission
from app.models.user import User
from app.models.item import ItemGrade, ItemName, ItemLength, ItemSize
from app.models.delivery import DeliveryTerm, DeliveryMode
from app.models.customer_classification import CustomerClassification
from app.models.contact_type import ContactType
from app.models.cost_point import CostPointMaster
from app.models.terms_condition import TermsNConditionMaster
from app.models.raw_material_cost import RawMaterialCost
from app.models.raw_material_cost_log import RawMaterialCostLog
from app.core.timezone import now_ist
from app.models.location import Country, StateMaster, DistrictMaster
from app.models.dia import DiaMaster
from app.models.status import EnQStatusMaster, QuotQStatusMaster
from app.models.communication import CommunicationMode
from app.models.financial_year import FinancialYear

router = APIRouter()


# ========== Schemas ==========

class ItemGradeCreate(BaseModel):
    itemGradeName: str

class ItemGradeResponse(BaseModel):
    itemGradeId: int
    companyId: int
    itemGradeName: str
    isActive: bool
    class Config:
        from_attributes = True

class ItemNameCreate(BaseModel):
    itemGradeId: int
    itemName: str
    itemDia: Optional[str] = None
    itemLength: Optional[str] = None
    erpItemCode: Optional[str] = None
    erpName: Optional[str] = None

class ItemNameResponse(BaseModel):
    itemId: int
    companyId: int
    itemGradeId: int
    itemGradeName: Optional[str] = None
    itemName: str
    itemDia: Optional[str] = None
    itemLength: Optional[str] = None
    erpItemCode: Optional[str] = None
    erpName: Optional[str] = None
    isActive: bool
    class Config:
        from_attributes = True

class ItemLengthCreate(BaseModel):
    itemId: int
    itemLength: str

class ItemLengthResponse(BaseModel):
    itemLengthId: int
    companyId: int
    itemId: int
    itemName: Optional[str] = None
    itemLength: str
    isActive: bool
    class Config:
        from_attributes = True

class ItemSizeCreate(BaseModel):
    itemId: int
    itemSize: str

class ItemSizeResponse(BaseModel):
    itemSizeId: int
    companyId: int
    itemId: int
    itemSize: str
    isActive: bool
    class Config:
        from_attributes = True

class DeliveryTermCreate(BaseModel):
    deliveryTerm: str

class DeliveryTermResponse(BaseModel):
    deliveryTermId: int
    companyId: int
    deliveryTerm: str
    isActive: bool
    class Config:
        from_attributes = True

class DeliveryModeCreate(BaseModel):
    deliveryMode: str

class DeliveryModeResponse(BaseModel):
    deliveryModeId: int
    companyId: int
    deliveryMode: str
    isActive: bool
    class Config:
        from_attributes = True

class ContactTypeCreate(BaseModel):
    contactType: str

class ContactTypeResponse(BaseModel):
    contactTypeId: int
    companyId: int
    contactType: str
    isActive: bool
    class Config:
        from_attributes = True

class ClassificationCreate(BaseModel):
    classificationName: str

class ClassificationResponse(BaseModel):
    classificationId: int
    companyId: int
    classificationName: str
    isActive: bool
    class Config:
        from_attributes = True

class CostPointCreate(BaseModel):
    costPointName: str
    isPrimary: bool = False
    isTax: bool = False

class CostPointResponse(BaseModel):
    costPointId: int
    companyId: int
    costPointName: str
    isPrimary: bool
    isTax: bool
    isActive: bool
    class Config:
        from_attributes = True

class TncCreate(BaseModel):
    tncName: str
    tncDescription: Optional[str] = None

class TncResponse(BaseModel):
    tncId: int
    companyId: int
    tncName: str
    tncDescription: Optional[str] = None
    isActive: bool
    class Config:
        from_attributes = True

class RawMaterialCostCreate(BaseModel):
    dia: str
    tpcost: float
    effectedFrom: Optional[datetime] = None
    isBasePrice: Optional[bool] = False
    diffFromBase: Optional[float] = None

class RawMaterialCostResponse(BaseModel):
    rawMaterialCostId: int
    companyId: int
    dia: str
    tpcost: float
    effectedFrom: Optional[datetime] = None
    isBasePrice: bool = False
    diffFromBase: Optional[float] = None
    createdbyName: Optional[str] = None
    createdon: Optional[datetime] = None
    lastupdatebyName: Optional[str] = None
    lastupdateon: Optional[datetime] = None
    isActive: bool
    class Config:
        from_attributes = True


class CountryCreate(BaseModel):
    countryname: str

class CountryResponse(BaseModel):
    countryid: int
    countryname: str
    isActive: bool
    class Config:
        from_attributes = True

class StateCreate(BaseModel):
    StateName: str
    Country: Optional[str] = "India"

class StateResponse(BaseModel):
    stateid: int
    StateName: str
    Country: Optional[str] = None
    isActive: bool
    class Config:
        from_attributes = True

class DistrictCreate(BaseModel):
    districName: str
    StateName: Optional[str] = None
    Country: Optional[str] = "India"

class DistrictResponse(BaseModel):
    districtid: int
    districName: str
    StateName: Optional[str] = None
    Country: Optional[str] = None
    isActive: bool
    class Config:
        from_attributes = True

class DiaMasterCreate(BaseModel):
    itemid: int
    diadescription: str

class DiaMasterResponse(BaseModel):
    diaid: int
    companyId: int
    itemid: int
    diadescription: str
    isActive: bool
    class Config:
        from_attributes = True

class EnQStatusCreate(BaseModel):
    enqStatus: str
    stepno: Optional[int] = None

class EnQStatusResponse(BaseModel):
    enqstatid: int
    companyId: int
    enqStatus: str
    stepno: Optional[int] = None
    isActive: bool
    class Config:
        from_attributes = True

class QuotQStatusCreate(BaseModel):
    quotStatus: str
    stepno: Optional[int] = None

class QuotQStatusResponse(BaseModel):
    quotstatid: int
    companyId: int
    quotStatus: str
    stepno: Optional[int] = None
    isActive: bool
    class Config:
        from_attributes = True

class CommunicationModeCreate(BaseModel):
    commmode: str

class CommunicationModeResponse(BaseModel):
    commmodeId: int
    companyId: int
    commmode: str
    isActive: bool
    class Config:
        from_attributes = True

class FinancialYearCreate(BaseModel):
    fyName: str
    fyCode: str
    isCurrent: Optional[bool] = False

class FinancialYearResponse(BaseModel):
    fyId: int
    companyId: int
    fyName: str
    fyCode: str
    isCurrent: bool
    isActive: bool
    class Config:
        from_attributes = True


# ========== Generic CRUD helper ==========

def _make_crud(prefix, model_cls, pk_field, create_schema, response_schema, menu_name):
    """Generate standard CRUD endpoints for a master table."""

    @router.get(f"/{prefix}", response_model=List[response_schema], tags=[menu_name])
    def list_items(
        db: Session = Depends(get_db),
        current_user: CurrentUser = Depends(get_current_user),
    ):
        return db.query(model_cls).filter(
            model_cls.companyId == current_user.company_id,
            model_cls.isActive == True,
        ).all()

    @router.post(f"/{prefix}", response_model=response_schema, status_code=201, tags=[menu_name])
    def create_item(
        data: create_schema,
        db: Session = Depends(get_db),
        current_user: CurrentUser = Depends(get_current_user),
    ):
        item = model_cls(
            **data.model_dump(),
            companyId=current_user.company_id,
            createdby=current_user.user_id,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return item

    @router.put(f"/{prefix}/{{item_id}}", response_model=response_schema, tags=[menu_name])
    def update_item(
        item_id: int,
        data: create_schema,
        db: Session = Depends(get_db),
        current_user: CurrentUser = Depends(get_current_user),
    ):
        pk_col = getattr(model_cls, pk_field)
        item = db.query(model_cls).filter(
            pk_col == item_id,
            model_cls.companyId == current_user.company_id,
            model_cls.isActive == True,
        ).first()
        if not item:
            raise HTTPException(status_code=404, detail="Not found")
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(item, key, value)
        item.lastupdateby = current_user.user_id
        db.commit()
        db.refresh(item)
        return item

    @router.delete(f"/{prefix}/{{item_id}}", status_code=204, tags=[menu_name])
    def delete_item(
        item_id: int,
        db: Session = Depends(get_db),
        current_user: CurrentUser = Depends(get_current_user),
    ):
        pk_col = getattr(model_cls, pk_field)
        item = db.query(model_cls).filter(
            pk_col == item_id,
            model_cls.companyId == current_user.company_id,
            model_cls.isActive == True,
        ).first()
        if not item:
            raise HTTPException(status_code=404, detail="Not found")
        item.isActive = False
        item.lastupdateby = current_user.user_id
        db.commit()


# ========== Custom list endpoints (with related names) ==========

@router.get("/item-names", response_model=List[ItemNameResponse], tags=["Item Name"])
def list_item_names(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    rows = (
        db.query(ItemName, ItemGrade.itemGradeName)
        .join(ItemGrade, ItemName.itemGradeId == ItemGrade.itemGradeId)
        .filter(ItemName.companyId == current_user.company_id, ItemName.isActive == True)
        .all()
    )
    result = []
    for item, grade_name in rows:
        d = {c.key: getattr(item, c.key) for c in ItemName.__table__.columns}
        d["itemGradeName"] = grade_name
        result.append(d)
    return result


@router.get("/item-lengths", response_model=List[ItemLengthResponse], tags=["Item Length"])
def list_item_lengths(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    rows = (
        db.query(ItemLength, ItemName.itemName)
        .join(ItemName, ItemLength.itemId == ItemName.itemId)
        .filter(ItemLength.companyId == current_user.company_id, ItemLength.isActive == True)
        .all()
    )
    result = []
    for length, item_name in rows:
        d = {c.key: getattr(length, c.key) for c in ItemLength.__table__.columns}
        d["itemName"] = item_name
        result.append(d)
    return result


@router.get("/raw-material-costs", response_model=List[RawMaterialCostResponse], tags=["Raw Material Cost"])
def list_raw_material_costs(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    CreatedByUser = aliased(User)
    UpdatedByUser = aliased(User)
    rows = (
        db.query(
            RawMaterialCost,
            CreatedByUser.userName.label("createdbyName"),
            UpdatedByUser.userName.label("lastupdatebyName"),
        )
        .outerjoin(CreatedByUser, RawMaterialCost.createdby == CreatedByUser.userId)
        .outerjoin(UpdatedByUser, RawMaterialCost.lastupdateby == UpdatedByUser.userId)
        .filter(
            RawMaterialCost.companyId == current_user.company_id,
            RawMaterialCost.isActive == True,
        )
        .all()
    )
    result = []
    for rmc, created_name, updated_name in rows:
        d = {c.key: getattr(rmc, c.key) for c in RawMaterialCost.__table__.columns}
        d["createdbyName"] = created_name
        d["lastupdatebyName"] = updated_name
        result.append(d)
    return result


# ========== Raw Material Cost: Update Logs ==========

class RawMaterialCostLogResponse(BaseModel):
    logId: int
    rawMaterialCostId: int
    dia: str
    oldCost: Optional[float] = None
    newCost: float
    oldEffectedFrom: Optional[datetime] = None
    newEffectedFrom: Optional[datetime] = None
    action: str
    remarks: Optional[str] = None
    changedBy: Optional[int] = None
    changedByName: Optional[str] = None
    changedOn: datetime

    class Config:
        from_attributes = True


class RawMaterialCostUpdateWithRemarks(BaseModel):
    dia: str
    tpcost: float
    effectedFrom: Optional[datetime] = None
    isBasePrice: Optional[bool] = None
    diffFromBase: Optional[float] = None
    remarks: Optional[str] = None  # optional reason for the change


@router.put(
    "/raw-material-costs/{item_id}",
    response_model=RawMaterialCostResponse,
    tags=["Raw Material Cost"],
)
def update_raw_material_cost(
    item_id: int,
    data: RawMaterialCostUpdateWithRemarks,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Update Raw Material Cost — writes an audit log row capturing old→new."""
    item = db.query(RawMaterialCost).filter(
        RawMaterialCost.rawMaterialCostId == item_id,
        RawMaterialCost.companyId == current_user.company_id,
        RawMaterialCost.isActive == True,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")

    old_cost = float(item.tpcost) if item.tpcost is not None else None
    old_effected_from = item.effectedFrom

    changed_cost = (old_cost != float(data.tpcost))
    changed_effected_from = (old_effected_from != data.effectedFrom)
    changed_dia = (item.dia != data.dia)

    # Enforce: only one row per company can be isBasePrice
    if data.isBasePrice:
        db.query(RawMaterialCost).filter(
            RawMaterialCost.companyId == current_user.company_id,
            RawMaterialCost.rawMaterialCostId != item_id,
            RawMaterialCost.isBasePrice == True,
            RawMaterialCost.isActive == True,
        ).update({"isBasePrice": False})

    item.dia = data.dia
    item.tpcost = data.tpcost
    item.effectedFrom = data.effectedFrom
    if data.isBasePrice is not None:
        item.isBasePrice = data.isBasePrice
    if data.diffFromBase is not None:
        item.diffFromBase = data.diffFromBase
    elif data.isBasePrice:
        item.diffFromBase = None  # base row has no difference
    item.lastupdateby = current_user.user_id

    # Log only when meaningful fields actually change
    if changed_cost or changed_effected_from or changed_dia:
        log = RawMaterialCostLog(
            rawMaterialCostId=item.rawMaterialCostId,
            companyId=current_user.company_id,
            dia=data.dia,
            oldCost=old_cost,
            newCost=data.tpcost,
            oldEffectedFrom=old_effected_from,
            newEffectedFrom=data.effectedFrom,
            action="UPDATE",
            remarks=data.remarks,
            changedBy=current_user.user_id,
            changedOn=now_ist(),
        )
        db.add(log)

    # CASCADE: If base price changed, recalc all dependent rows
    if data.isBasePrice and changed_cost:
        new_base = float(data.tpcost)
        dependent_rows = db.query(RawMaterialCost).filter(
            RawMaterialCost.companyId == current_user.company_id,
            RawMaterialCost.rawMaterialCostId != item_id,
            RawMaterialCost.isBasePrice == False,
            RawMaterialCost.isActive == True,
        ).all()
        for dep in dependent_rows:
            dep_old_cost = float(dep.tpcost) if dep.tpcost is not None else 0
            dep_diff = float(dep.diffFromBase) if dep.diffFromBase is not None else 0
            dep_new_cost = round(new_base + dep_diff, 2)
            dep_old_effected = dep.effectedFrom
            # Update cost and effectedFrom to match base
            dep.tpcost = dep_new_cost
            dep.effectedFrom = data.effectedFrom
            dep.lastupdateby = current_user.user_id
            # Log the cascade change
            if dep_old_cost != dep_new_cost or dep_old_effected != data.effectedFrom:
                db.add(RawMaterialCostLog(
                    rawMaterialCostId=dep.rawMaterialCostId,
                    companyId=current_user.company_id,
                    dia=dep.dia,
                    oldCost=dep_old_cost,
                    newCost=dep_new_cost,
                    oldEffectedFrom=dep_old_effected,
                    newEffectedFrom=data.effectedFrom,
                    action="CASCADE",
                    remarks=f"Base price changed: {old_cost} → {data.tpcost}",
                    changedBy=current_user.user_id,
                    changedOn=now_ist(),
                ))

    db.commit()
    db.refresh(item)

    # If effectedFrom is today, trigger immediate TP cost update for Draft quotations
    today = now_ist().date()
    eff_date = data.effectedFrom.date() if isinstance(data.effectedFrom, datetime) else data.effectedFrom
    if eff_date and eff_date <= today:
        from app.services.tp_cost_background import trigger_immediate_update
        trigger_immediate_update(current_user.company_id)

    return item


@router.post(
    "/raw-material-costs/apply-tp-update",
    tags=["Raw Material Cost"],
)
def apply_tp_cost_update(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Manually trigger TP cost update for Draft quotations in this company.
    Updates TPWGST based on latest RawMaterialCost where effectedFrom <= today.
    """
    from app.services.tp_cost_scheduler import run_tp_cost_update
    results = run_tp_cost_update(db, company_id=current_user.company_id)
    company_result = results.get(current_user.company_id, {})
    return {
        "message": "TP cost update applied",
        "quotationDetailsUpdated": company_result.get("quot_details", 0),
    }


@router.get(
    "/raw-material-costs/base-price",
    tags=["Raw Material Cost"],
)
def get_base_price(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Return the base-price row for this company (isBasePrice=True), or null."""
    row = db.query(RawMaterialCost).filter(
        RawMaterialCost.companyId == current_user.company_id,
        RawMaterialCost.isBasePrice == True,
        RawMaterialCost.isActive == True,
    ).first()
    if not row:
        return None
    return {
        "rawMaterialCostId": row.rawMaterialCostId,
        "dia": row.dia,
        "tpcost": float(row.tpcost),
    }


@router.get(
    "/raw-material-costs/{item_id}/logs",
    tags=["Raw Material Cost"],
)
def list_raw_material_cost_logs(
    item_id: int,
    dateFrom: Optional[datetime] = None,
    dateTo: Optional[datetime] = None,
    page: int = 1,
    pageSize: int = 25,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Paginated update history for a single Raw Material Cost row.

    Query params:
      - dateFrom, dateTo: optional ISO datetime range (by changedOn)
      - page (1-based), pageSize (default 25, max 200)
    """
    # Bound pageSize
    pageSize = max(1, min(200, pageSize))
    page = max(1, page)

    # Verify the parent belongs to the user's company
    parent = db.query(RawMaterialCost).filter(
        RawMaterialCost.rawMaterialCostId == item_id,
        RawMaterialCost.companyId == current_user.company_id,
    ).first()
    if not parent:
        raise HTTPException(status_code=404, detail="Raw Material Cost not found")

    ChangedByUser = aliased(User)
    q = (
        db.query(RawMaterialCostLog, ChangedByUser.userName.label("changedByName"))
        .outerjoin(ChangedByUser, RawMaterialCostLog.changedBy == ChangedByUser.userId)
        .filter(
            RawMaterialCostLog.rawMaterialCostId == item_id,
            RawMaterialCostLog.companyId == current_user.company_id,
        )
    )
    if dateFrom:
        q = q.filter(RawMaterialCostLog.changedOn >= dateFrom)
    if dateTo:
        q = q.filter(RawMaterialCostLog.changedOn <= dateTo)

    total = q.count()
    q = q.order_by(RawMaterialCostLog.changedOn.desc()).offset((page - 1) * pageSize).limit(pageSize)

    items = []
    for log, changed_name in q.all():
        items.append({
            "logId": log.logId,
            "rawMaterialCostId": log.rawMaterialCostId,
            "dia": log.dia,
            "oldCost": float(log.oldCost) if log.oldCost is not None else None,
            "newCost": float(log.newCost),
            "oldEffectedFrom": log.oldEffectedFrom,
            "newEffectedFrom": log.newEffectedFrom,
            "action": log.action,
            "remarks": log.remarks,
            "changedBy": log.changedBy,
            "changedByName": changed_name,
            "changedOn": log.changedOn,
        })
    return {
        "items": items,
        "total": total,
        "page": page,
        "pageSize": pageSize,
        "totalPages": (total + pageSize - 1) // pageSize,
    }


@router.get(
    "/raw-material-costs/{item_id}/logs/export-excel",
    tags=["Raw Material Cost"],
)
def export_raw_material_cost_logs_excel(
    item_id: int,
    dateFrom: Optional[datetime] = None,
    dateTo: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Download all update logs (within date range) as XLSX."""
    from fastapi.responses import Response
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    parent = db.query(RawMaterialCost).filter(
        RawMaterialCost.rawMaterialCostId == item_id,
        RawMaterialCost.companyId == current_user.company_id,
    ).first()
    if not parent:
        raise HTTPException(status_code=404, detail="Raw Material Cost not found")

    ChangedByUser = aliased(User)
    q = (
        db.query(RawMaterialCostLog, ChangedByUser.userName.label("changedByName"))
        .outerjoin(ChangedByUser, RawMaterialCostLog.changedBy == ChangedByUser.userId)
        .filter(
            RawMaterialCostLog.rawMaterialCostId == item_id,
            RawMaterialCostLog.companyId == current_user.company_id,
        )
    )
    if dateFrom:
        q = q.filter(RawMaterialCostLog.changedOn >= dateFrom)
    if dateTo:
        q = q.filter(RawMaterialCostLog.changedOn <= dateTo)
    logs = q.order_by(RawMaterialCostLog.changedOn.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Update Logs"

    # Header context rows
    ws.cell(row=1, column=1, value="Raw Material Cost Update Logs").font = Font(size=14, bold=True)
    ws.cell(row=2, column=1, value="Dia:").font = Font(bold=True)
    ws.cell(row=2, column=2, value=parent.dia)
    ws.cell(row=3, column=1, value="Current Cost:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=float(parent.tpcost) if parent.tpcost else 0)
    ws.cell(row=4, column=1, value="Date Range:").font = Font(bold=True)
    range_str = ""
    if dateFrom:
        range_str += f"From {dateFrom.strftime('%d-%b-%Y %H:%M')} "
    if dateTo:
        range_str += f"To {dateTo.strftime('%d-%b-%Y %H:%M')}"
    if not range_str:
        range_str = "All"
    ws.cell(row=4, column=2, value=range_str)

    # Column headers
    header_row = 6
    headers = [
        ("Log ID", 8),
        ("Dia", 10),
        ("Action", 10),
        ("Old Cost", 12),
        ("New Cost", 12),
        ("Change", 12),
        ("Old Effected From", 18),
        ("New Effected From", 18),
        ("Remarks", 30),
        ("Changed By", 20),
        ("Changed On (IST)", 20),
    ]
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    for col_idx, (label, width) in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, (log, changed_name) in enumerate(logs, start=header_row + 1):
        old_cost = float(log.oldCost) if log.oldCost is not None else None
        new_cost = float(log.newCost)
        delta = (new_cost - old_cost) if old_cost is not None else None
        values = [
            log.logId,
            log.dia,
            log.action,
            old_cost,
            new_cost,
            delta,
            log.oldEffectedFrom.strftime("%d-%b-%Y %H:%M") if log.oldEffectedFrom else "",
            log.newEffectedFrom.strftime("%d-%b-%Y %H:%M") if log.newEffectedFrom else "",
            log.remarks or "",
            changed_name or "",
            log.changedOn.strftime("%d-%b-%Y %H:%M") if log.changedOn else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.border = border
            if col_idx in (4, 5, 6) and isinstance(value, (int, float)):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
                # Red for price drops (delta < 0) on the Change column
                if col_idx == 6 and value is not None and value < 0:
                    c.font = Font(color="C00000")
                elif col_idx == 6 and value is not None and value > 0:
                    c.font = Font(color="00703C")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"rawmat-cost-{parent.dia}-logs.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ========== Register all masters ==========

_make_crud("item-grades", ItemGrade, "itemGradeId", ItemGradeCreate, ItemGradeResponse, "Item Grade")
_make_crud("item-names", ItemName, "itemId", ItemNameCreate, ItemNameResponse, "Item Name")
_make_crud("item-lengths", ItemLength, "itemLengthId", ItemLengthCreate, ItemLengthResponse, "Item Length")
_make_crud("item-sizes", ItemSize, "itemSizeId", ItemSizeCreate, ItemSizeResponse, "Item Size")
_make_crud("delivery-terms", DeliveryTerm, "deliveryTermId", DeliveryTermCreate, DeliveryTermResponse, "Delivery Term")
_make_crud("delivery-modes", DeliveryMode, "deliveryModeId", DeliveryModeCreate, DeliveryModeResponse, "Delivery Mode")
_make_crud("contact-types", ContactType, "contactTypeId", ContactTypeCreate, ContactTypeResponse, "Contact Type")
_make_crud("customer-classifications", CustomerClassification, "classificationId", ClassificationCreate, ClassificationResponse, "Customer Classification")
_make_crud("cost-points", CostPointMaster, "costPointId", CostPointCreate, CostPointResponse, "Cost Point")
_make_crud("terms-conditions", TermsNConditionMaster, "tncId", TncCreate, TncResponse, "Terms & Conditions")
# Custom POST for raw-material-costs (trigger TP update on create if effectedFrom <= today)
@router.post("/raw-material-costs", response_model=RawMaterialCostResponse, status_code=201, tags=["Raw Material Cost"])
def create_raw_material_cost(
    data: RawMaterialCostCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    if data.isBasePrice:
        db.query(RawMaterialCost).filter(
            RawMaterialCost.companyId == current_user.company_id,
            RawMaterialCost.isBasePrice == True,
            RawMaterialCost.isActive == True,
        ).update({"isBasePrice": False})

    item = RawMaterialCost(
        **data.model_dump(),
        companyId=current_user.company_id,
        createdby=current_user.user_id,
    )
    db.add(item)
    db.commit()
    db.refresh(item)

    # Trigger immediate TP update if effectedFrom <= today
    today = now_ist().date()
    eff = data.effectedFrom.date() if isinstance(data.effectedFrom, datetime) else data.effectedFrom
    if eff and eff <= today:
        from app.services.tp_cost_background import trigger_immediate_update
        trigger_immediate_update(current_user.company_id)

    return item

_make_crud("raw-material-costs", RawMaterialCost, "rawMaterialCostId", RawMaterialCostCreate, RawMaterialCostResponse, "Raw Material Cost")
_make_crud("dia-masters", DiaMaster, "diaid", DiaMasterCreate, DiaMasterResponse, "Dia Master")
_make_crud("enq-statuses", EnQStatusMaster, "enqstatid", EnQStatusCreate, EnQStatusResponse, "Enquiry Status")
_make_crud("quot-statuses", QuotQStatusMaster, "quotstatid", QuotQStatusCreate, QuotQStatusResponse, "Quotation Status")
_make_crud("communication-modes", CommunicationMode, "commmodeId", CommunicationModeCreate, CommunicationModeResponse, "Communication Mode")
_make_crud("financial-years", FinancialYear, "fyId", FinancialYearCreate, FinancialYearResponse, "Financial Year")


# ========== Location masters (global, no companyId) ==========

@router.get("/countries", response_model=List[CountryResponse], tags=["Country"])
def list_countries(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    return db.query(Country).filter(Country.isActive == True).all()


@router.post("/countries", response_model=CountryResponse, status_code=201, tags=["Country"])
def create_country(
    data: CountryCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = Country(**data.model_dump(), createdby=current_user.user_id)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/countries/{item_id}", response_model=CountryResponse, tags=["Country"])
def update_country(
    item_id: int,
    data: CountryCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = db.query(Country).filter(Country.countryid == item_id, Country.isActive == True).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    item.lastupdateby = current_user.user_id
    db.commit()
    db.refresh(item)
    return item


@router.delete("/countries/{item_id}", status_code=204, tags=["Country"])
def delete_country(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = db.query(Country).filter(Country.countryid == item_id, Country.isActive == True).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    item.isActive = False
    item.lastupdateby = current_user.user_id
    db.commit()


@router.get("/states", response_model=List[StateResponse], tags=["State"])
def list_states(
    country: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    q = db.query(StateMaster).filter(StateMaster.isActive == True)
    if country:
        q = q.filter(StateMaster.Country == country)
    return q.order_by(StateMaster.StateName).all()


@router.post("/states", response_model=StateResponse, status_code=201, tags=["State"])
def create_state(
    data: StateCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = StateMaster(**data.model_dump(), createdby=current_user.user_id)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/states/{item_id}", response_model=StateResponse, tags=["State"])
def update_state(
    item_id: int,
    data: StateCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = db.query(StateMaster).filter(StateMaster.stateid == item_id, StateMaster.isActive == True).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    item.lastupdateby = current_user.user_id
    db.commit()
    db.refresh(item)
    return item


@router.delete("/states/{item_id}", status_code=204, tags=["State"])
def delete_state(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = db.query(StateMaster).filter(StateMaster.stateid == item_id, StateMaster.isActive == True).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    item.isActive = False
    item.lastupdateby = current_user.user_id
    db.commit()


@router.get("/districts", response_model=List[DistrictResponse], tags=["District"])
def list_districts(
    state: Optional[str] = None,
    country: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    q = db.query(DistrictMaster).filter(DistrictMaster.isActive == True)
    if state:
        q = q.filter(DistrictMaster.StateName == state)
    if country:
        q = q.filter(DistrictMaster.Country == country)
    return q.order_by(DistrictMaster.districName).all()


@router.post("/districts", response_model=DistrictResponse, status_code=201, tags=["District"])
def create_district(
    data: DistrictCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = DistrictMaster(**data.model_dump(), createdby=current_user.user_id)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/districts/{item_id}", response_model=DistrictResponse, tags=["District"])
def update_district(
    item_id: int,
    data: DistrictCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = db.query(DistrictMaster).filter(DistrictMaster.districtid == item_id, DistrictMaster.isActive == True).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    item.lastupdateby = current_user.user_id
    db.commit()
    db.refresh(item)
    return item


@router.delete("/districts/{item_id}", status_code=204, tags=["District"])
def delete_district(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    item = db.query(DistrictMaster).filter(DistrictMaster.districtid == item_id, DistrictMaster.isActive == True).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    item.isActive = False
    item.lastupdateby = current_user.user_id
    db.commit()
