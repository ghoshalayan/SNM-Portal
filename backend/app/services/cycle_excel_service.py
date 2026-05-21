"""Cycle-scoped XLSX export (Phase 1G).

Produces one workbook per cycle with the following sheets:

  1. **Summary**          — cycle envelope: cycleNo, status, started /
                             closed metadata, parent cycle, notes, plus
                             a compact roster of every PO/LOI in the
                             cycle.
  2. **Working Sheet**    — every active ``QuotPOWorkingSheet`` row
                             attached to the cycle, same column shape
                             as ``viability_excel_service.WORKING_COLUMNS``.
  3. **Viability Sheet**  — populated when the cycle has an Approved
                             viability; otherwise omitted.

The styling helpers are imported from ``viability_excel_service`` so
both exports look visually identical — the cycle Excel is just a
broader fan-out over the same per-line column shape.
"""
from __future__ import annotations

from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.quot_annexure import QuotAnnexure
from app.models.quot_order_cycle import QuotOrderCycle
from app.models.quot_po_working_sheet import QuotPOWorkingSheet
from app.models.quot_purchase_order import QuotPurchaseOrder
from app.models.quot_viability import QuotViabilityLine, QuotViabilitySheet
from app.models.quotation import QuotSummary
from app.services.viability_excel_service import (
    COL_HEADER_FONT,
    HEADER_FILL,
    HEADER_LABEL_FONT,
    HEADER_VALUE_FONT,
    THIN_BORDER,
    VIABILITY_COLUMNS,
    WORKING_COLUMNS,
    _write_column_headers,
    _write_data_rows,
)


def _write_summary_sheet(
    ws,
    *,
    quot: Optional[QuotSummary],
    cycle: QuotOrderCycle,
    parent_cycle: Optional[QuotOrderCycle],
    purchase_orders: List[QuotPurchaseOrder],
    viability: Optional[QuotViabilitySheet],
    annexure: Optional[QuotAnnexure],
) -> None:
    """Top-of-workbook cycle envelope + PO roster. Single sheet, no
    cost-head heavy lifting — just metadata the user wants at a
    glance when they pop the file open."""
    pairs = [
        ("Quotation No:", quot.quotNo if quot and quot.quotNo else ""),
        (
            "Customer:",
            (quot.customer.customerName if quot and quot.customer else "") or "",
        ),
        ("Cycle:", f"Cycle {cycle.cycleNo}"),
        ("Cycle Status:", cycle.status),
        (
            "Started On:",
            cycle.startedOn.strftime("%d-%b-%Y") if cycle.startedOn else "",
        ),
        (
            "Closed On:",
            cycle.closedOn.strftime("%d-%b-%Y") if cycle.closedOn else "",
        ),
        (
            "Parent Cycle:",
            f"Cycle {parent_cycle.cycleNo}" if parent_cycle else "—",
        ),
        ("Viability:", viability.status if viability else "—"),
        ("Annexure:", annexure.status if annexure else "—"),
        ("Notes:", (cycle.notes or "").replace("\n", " · ")),
    ]
    for row_idx, (label, value) in enumerate(pairs, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = HEADER_LABEL_FONT
        if value:
            ws.cell(row=row_idx, column=2, value=value).font = HEADER_VALUE_FONT

    # POs / LOIs roster two rows below the envelope.
    roster_header_row = len(pairs) + 2
    columns = [
        ("Seq", 6),
        ("Type", 8),
        ("PO No", 18),
        ("PO Date", 12),
        ("Status", 12),
        ("Remarks", 30),
    ]
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=roster_header_row, column=col_idx, value=label)
        cell.font = COL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_offset, po in enumerate(purchase_orders, start=1):
        row_idx = roster_header_row + row_offset
        values = [
            po.loiSequence,
            "LOI" if getattr(po, "isLOI", False) else "PO",
            po.poNo,
            po.poDate.strftime("%d-%b-%Y") if po.poDate else "",
            po.status,
            po.remarks or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx <= 5 else "left",
                vertical="center",
            )


def build_cycle_xlsx(
    *,
    quot: Optional[QuotSummary],
    cycle: QuotOrderCycle,
    parent_cycle: Optional[QuotOrderCycle],
    purchase_orders: List[QuotPurchaseOrder],
    working_sheet: List[QuotPOWorkingSheet],
    viability: Optional[QuotViabilitySheet],
    viability_lines: List[QuotViabilityLine],
    annexure: Optional[QuotAnnexure],
) -> bytes:
    """Build the full cycle workbook. The caller fetches everything;
    this function is pure — no DB access — so it stays trivial to test.

    The Viability sheet is omitted when ``viability`` is None (the
    cycle hasn't generated a viability yet); same for the rows. The
    Summary + Working Sheet are always present.
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(
        ws_summary,
        quot=quot, cycle=cycle, parent_cycle=parent_cycle,
        purchase_orders=purchase_orders,
        viability=viability, annexure=annexure,
    )

    # Sheet 2: Working Sheet (always rendered, even when empty)
    ws_working = wb.create_sheet(title="Working Sheet")
    header_row = _emit_quot_header(ws_working, quot, f"Cycle {cycle.cycleNo} — Working Sheet")
    _write_column_headers(ws_working, WORKING_COLUMNS, header_row)
    _write_data_rows(ws_working, WORKING_COLUMNS, working_sheet, header_row)

    # Sheet 3: Viability Sheet — only when the cycle has one
    if viability is not None:
        ws_viab = wb.create_sheet(title="Viability Sheet")
        header_row = _emit_quot_header(ws_viab, quot, f"Cycle {cycle.cycleNo} — Viability Sheet")
        _write_column_headers(ws_viab, VIABILITY_COLUMNS, header_row)
        _write_data_rows(ws_viab, VIABILITY_COLUMNS, viability_lines, header_row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _emit_quot_header(ws, quot: Optional[QuotSummary], sheet_title: str) -> int:
    """Light variant of ``viability_excel_service._write_header_rows``
    — same shape so the workbook reads consistently, but we control
    the sheet label here so the cycle number lands in the header."""
    client_name = ""
    site_name = ""
    quot_no = ""
    quot_date = ""
    if quot is not None:
        if quot.customer is not None:
            client_name = quot.customer.customerName or ""
        if quot.site is not None:
            site_name = (
                quot.site.siteAddressCode
                or getattr(quot.site, "addressLine", "")
                or ""
            )
        quot_no = quot.quotNo or ""
        quot_date = quot.quotDate.strftime("%d-%b-%Y") if quot.quotDate else ""

    pairs = [
        ("Quotation No:", quot_no),
        ("Client name:", client_name),
        ("Site Name:", site_name),
        ("Date:", quot_date),
        ("Sheet:", sheet_title),
    ]
    for idx, (label, value) in enumerate(pairs, start=1):
        ws.cell(row=idx, column=1, value=label).font = HEADER_LABEL_FONT
        if value:
            ws.cell(row=idx, column=2, value=value).font = HEADER_VALUE_FONT
    return len(pairs) + 2
