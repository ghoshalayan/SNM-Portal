"""Generate Excel export for Quotation Line Items.

Produces an XLSX file matching the standard quotation working format:
- Header rows (Client name, Site Name, Payment terms, T.P. Ref., Date)
- Column headers (Sl no., Item, Dia, Length, Qty, T.P. w/o GST, Marketing,
  Freight Trailer, Freight Truck, Unloading, OHD, IFC, Weighment Diff., CD,
  S&E Charges, CRS, Incidental, Short Length, Specific Length, Extra,
  Fluctuation, Commission, Misc., Testing, MOU TOD, Special Discount, JC,
  Total (Rs/MT), GST @ 18%, EX/FOR Price)
- Data rows with red font for negative-charge columns (CD, Short Length, Special Discount)
- Yellow highlight on rows matching T.P. Ref dia (16mm by default)
"""

from io import BytesIO
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Column order MUST match the standard format exactly
COLUMNS = [
    ("Sl no.", "_sno", 6),
    ("Item", "itemName", 20),
    ("Grade", "itemGradeName", 16),
    ("Dia (mm)", "itemDia", 10),
    ("Length", "itemLength", 18),
    ("Qty (MT)", "quantity", 9),
    ("T.P. w/o GST", "TPWGST", 12),
    ("Marketing", "Marketing", 11),
    ("Freight Trailer", "FreightTrailer", 12),
    ("Freight Truck", "FreightTruck", 12),
    ("Unloading", "Unloading", 11),
    ("OHD", "OHD", 8),
    ("IFC", "IFC", 8),
    ("Weighment Diff.", "WeighmentDiff", 13),
    ("CD", "CD", 8),
    ("S & E Charges", "SWECharge", 13),
    ("CRS", "CRS", 8),
    ("Incidental Charges", "IncCharge", 13),
    ("Short Length Charges", "ShortLnthCharge", 14),
    ("Specific Length Charges", "SpeciFicLnthCharge", 14),
    ("Extra", "ExtraCharge", 9),
    ("Fluctuation", "Fluctuation", 11),
    ("Commission", "Commission", 11),
    ("Misc.", "Misc", 8),
    ("Testing", "Testing", 9),
    ("MOU TOD", "MOUTOD", 10),
    ("Special Discount", "SplDisc", 13),
    ("JC", "JC", 8),
    ("Total (Rs/MT)", "totRate", 13),
    ("GST @ 18%", "_gst", 11),
    ("EX/FOR Price", "totAmount", 13),
    ("Mode of Dispatch", "modeOfDispatch", 22),
]

# Cost-head columns that conventionally hold negative values (shown red)
NEGATIVE_COLS = {"CD", "ShortLnthCharge", "SplDisc"}

# Style definitions
HEADER_LABEL_FONT = Font(name="Calibri", size=10, bold=True)
HEADER_VALUE_FONT = Font(name="Calibri", size=10)
COL_HEADER_FONT = Font(name="Calibri", size=10, bold=True)
NEG_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="C00000")
DATA_FONT = Font(name="Calibri", size=10)
NEG_DATA_FONT = Font(name="Calibri", size=10, color="C00000")
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def build_quotation_xlsx(
    *,
    client_name: str = "",
    site_name: str = "",
    payment_terms: str = "",
    tp_ref_dia: str = "16",
    quot_date: Optional[str] = None,
    quot_no: str = "",
    details: List[dict],
) -> bytes:
    """Build the XLSX bytes for a quotation's line items.

    `details` is a list of dicts with keys matching QuotDetails columns
    (itemGradeName, itemDia, itemLength, itemUnit, quantity, TPWGST, ... totAmount).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation Line Items"

    # ---- Top header rows (left-aligned key/value pairs) ----
    header_pairs = [
        ("Client name:", client_name),
        ("Site Name:", site_name),
        ("Payment terms:", payment_terms),
        (f"T.P. Ref.: ({tp_ref_dia} mm dia rate)", ""),
        ("Date:", quot_date or ""),
    ]
    if quot_no:
        header_pairs.insert(0, ("Quotation No:", quot_no))

    for row_idx, (label, value) in enumerate(header_pairs, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = HEADER_LABEL_FONT
        if value:
            ws.cell(row=row_idx, column=2, value=value).font = HEADER_VALUE_FONT

    # ---- Column headers ----
    header_row = len(header_pairs) + 2
    for col_idx, (label, key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = NEG_HEADER_FONT if key in NEGATIVE_COLS else COL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Set header row height for wrapped text
    ws.row_dimensions[header_row].height = 32

    # ---- Data rows ----
    data_start_row = header_row + 1
    for row_idx, detail in enumerate(details, start=data_start_row):
        sno = row_idx - data_start_row + 1
        # Highlight rows matching T.P. Ref dia (e.g. 16 mm)
        is_highlight = False
        try:
            dia_val = str(detail.get("itemDia", "")).strip()
            if dia_val == str(tp_ref_dia).strip():
                is_highlight = True
        except Exception:
            pass

        for col_idx, (label, key, _w) in enumerate(COLUMNS, start=1):
            if key == "_sno":
                value = sno
            elif key == "_gst":
                # GST = totRate * 0.18 (matches frontend logic)
                tot = detail.get("totRate") or 0
                try:
                    value = round(float(tot) * 0.18, 2)
                except Exception:
                    value = 0
            else:
                value = detail.get(key)
                if value is None or value == "":
                    value = "" if key in ("itemGradeName", "itemDia", "itemLength") else 0

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            # Red font for negative-convention columns
            cell.font = NEG_DATA_FONT if key in NEGATIVE_COLS else DATA_FONT

            # Yellow highlight for matching T.P. Ref dia row
            if is_highlight:
                cell.fill = HIGHLIGHT_FILL

            # Number formatting
            if isinstance(value, (int, float)) and key not in ("_sno",):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif key == "_sno":
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Freeze panes below header row
    ws.freeze_panes = ws.cell(row=data_start_row, column=4)

    # Serialize to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
