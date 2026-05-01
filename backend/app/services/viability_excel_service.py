"""Viability Sheet XLSX export.

Produces a single workbook with two sheets:
  1. "Working Sheet"    — snapshot of QuotDetails (columns identical to the
                           existing quotation Excel).
  2. "Viability Sheet"  — QuotViabilityLine rows with 4 extra columns after
                           EX/FOR Price: Ordered QTY, Total Amount, Total GST,
                           Gross EX/FOR Price.

Styling mirrors app/services/quotation_excel_service.py so the two exports
look consistent.
"""
from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.quotation import QuotSummary
from app.models.quot_viability import QuotViabilityLine


# Column definitions mirror quotation_excel_service.COLUMNS so the working
# sheet tab is visually identical to the standalone quotation export.
WORKING_COLUMNS = [
    ("Sl no.", "_sno", 6),
    ("Item", "itemName", 20),
    ("Grade", "itemGradeName", 16),
    ("Dia (mm)", "itemDia", 10),
    ("Length", "itemLength", 18),
    ("Qty (MT)", "quantity", 9),
    ("T.P. w/o GST", "TPWGST", 12),
    ("Marketing", "Marketing", 11),
    ("Freight Trailer", "FreightTrailer", 12),
    ("Freight Truck", "FreightTruck", 12),
    ("Unloading", "Unloading", 11),
    ("OHD", "OHD", 8),
    ("IFC", "IFC", 8),
    ("Weighment Diff.", "WeighmentDiff", 13),
    ("CD", "CD", 8),
    ("S & E Charges", "SWECharge", 13),
    ("CRS", "CRS", 8),
    ("Incidental Charges", "IncCharge", 13),
    ("Short Length Charges", "ShortLnthCharge", 14),
    ("Specific Length Charges", "SpeciFicLnthCharge", 14),
    ("Extra", "ExtraCharge", 9),
    ("Fluctuation", "Fluctuation", 11),
    ("Commission", "Commission", 11),
    ("Misc.", "Misc", 8),
    ("Testing", "Testing", 9),
    ("MOU TOD", "MOUTOD", 10),
    ("Special Discount", "SplDisc", 13),
    ("JC", "JC", 8),
    ("Total (Rs/MT)", "totRate", 13),
    ("GST @ 18%", "_gst", 11),
    ("EX/FOR Price", "totAmount", 13),
    ("Mode of Dispatch", "modeOfDispatch", 22),
]

# Viability adds 4 columns after EX/FOR Price
VIABILITY_EXTRA_COLS = [
    ("Ordered Qty (MT)", "orderedQty", 12),
    ("Total Amount (Rs)", "totalAmount", 15),
    ("Total GST", "totalGst", 13),
    ("Gross EX/FOR Price", "grossExForPrice", 16),
]

VIABILITY_COLUMNS = (
    WORKING_COLUMNS[:-1]  # everything except the last (Mode of Dispatch)
    + VIABILITY_EXTRA_COLS
    + [WORKING_COLUMNS[-1]]  # Mode of Dispatch at the end
)

NEGATIVE_COLS = {"CD", "ShortLnthCharge", "SplDisc"}

# Styles
HEADER_LABEL_FONT = Font(name="Calibri", size=10, bold=True)
HEADER_VALUE_FONT = Font(name="Calibri", size=10)
COL_HEADER_FONT = Font(name="Calibri", size=10, bold=True)
NEG_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="C00000")
DATA_FONT = Font(name="Calibri", size=10)
NEG_DATA_FONT = Font(name="Calibri", size=10, color="C00000")
HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GROSS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _row_value(obj, key: str):
    """Safe attribute lookup for both ORM objects and pydantic models."""
    if obj is None:
        return None
    if hasattr(obj, key):
        return getattr(obj, key)
    if isinstance(obj, dict):
        return obj.get(key)
    return None


def _write_header_rows(ws, quot: Optional[QuotSummary], sheet_title: str) -> int:
    """Writes top header (quot meta + sheet label) and returns the row index
    where the column-header row should go.
    """
    client_name = ""
    site_name = ""
    quot_no = ""
    quot_date = ""
    if quot is not None:
        if quot.customer is not None:
            client_name = quot.customer.customerName or ""
        if quot.site is not None:
            site_name = quot.site.siteAddressCode or getattr(quot.site, "addressLine", "") or ""
        quot_no = quot.quotNo or ""
        quot_date = quot.quotDate.strftime("%d-%b-%Y") if quot.quotDate else ""

    pairs = [
        ("Quotation No:", quot_no),
        ("Client name:", client_name),
        ("Site Name:", site_name),
        ("Date:", quot_date),
        ("Sheet:", sheet_title),
    ]
    for idx, (label, value) in enumerate(pairs, start=1):
        ws.cell(row=idx, column=1, value=label).font = HEADER_LABEL_FONT
        if value:
            ws.cell(row=idx, column=2, value=value).font = HEADER_VALUE_FONT
    return len(pairs) + 2


def _write_column_headers(ws, columns, header_row: int) -> None:
    for col_idx, (label, key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = NEG_HEADER_FONT if key in NEGATIVE_COLS else COL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 32


def _write_data_rows(ws, columns, rows, header_row: int) -> None:
    gross_keys = {k for _, k, _ in VIABILITY_EXTRA_COLS}
    data_start = header_row + 1
    for row_idx, line in enumerate(rows, start=data_start):
        sno = row_idx - data_start + 1
        for col_idx, (_label, key, _w) in enumerate(columns, start=1):
            if key == "_sno":
                value = sno
            elif key == "_gst":
                tot = _row_value(line, "totRate") or 0
                try:
                    value = round(float(tot) * 0.18, 2)
                except Exception:
                    value = 0
            else:
                value = _row_value(line, key)
                if value is None or value == "":
                    value = "" if key in ("itemGradeName", "itemDia", "itemLength", "modeOfDispatch") else 0

            cell = ws.cell(row=row_idx, column=col_idx, value=float(value) if isinstance(value, (int, float)) or (hasattr(value, "__float__") and not isinstance(value, str)) else value)
            cell.border = THIN_BORDER
            cell.font = NEG_DATA_FONT if key in NEGATIVE_COLS else DATA_FONT
            if key in gross_keys:
                cell.fill = GROSS_FILL

            if key == "_sno":
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Freeze panes below the column headers
    ws.freeze_panes = ws.cell(row=data_start, column=4)


def build_viability_xlsx(
    *,
    quot: Optional[QuotSummary],
    working_sheet: List,
    viability_lines: List[QuotViabilityLine],
) -> bytes:
    wb = Workbook()

    # ---- Sheet 1: Working Sheet ----
    ws_working = wb.active
    ws_working.title = "Working Sheet"
    header_row = _write_header_rows(ws_working, quot, "Working Sheet (original)")
    _write_column_headers(ws_working, WORKING_COLUMNS, header_row)
    _write_data_rows(ws_working, WORKING_COLUMNS, working_sheet, header_row)

    # ---- Sheet 2: Viability Sheet ----
    ws_viab = wb.create_sheet(title="Viability Sheet")
    header_row = _write_header_rows(ws_viab, quot, "Viability Sheet (adjusted)")
    _write_column_headers(ws_viab, VIABILITY_COLUMNS, header_row)
    _write_data_rows(ws_viab, VIABILITY_COLUMNS, viability_lines, header_row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
